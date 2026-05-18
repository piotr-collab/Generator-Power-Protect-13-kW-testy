from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path("/Users/piotrlunkiewicz/Documents/Warsztaty/PP_EcoGenerator_13kW_plan_badan.xlsx")
PHOTO = Path("/Users/piotrlunkiewicz/Documents/Codex/Power Protect/Power Protect Eco Generator 1-kopia.jpeg")

BLUE = "0B5E8E"
CYAN = "22A6C8"
LIGHT_BLUE = "D9EEF5"
PALE = "F6FAFC"
DARK = "1F2933"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
RED = "F4CCCC"
GREY = "E7EAED"
WHITE = "FFFFFF"


thin = Side(style="thin", color="B7C7D3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_title(ws, cell, title, subtitle=None):
    ws[cell] = title
    ws[cell].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws[cell].fill = PatternFill("solid", fgColor=BLUE)
    ws[cell].alignment = Alignment(vertical="center")
    ws.row_dimensions[ws[cell].row].height = 32
    if subtitle:
        sub = ws.cell(ws[cell].row + 1, ws[cell].column)
        sub.value = subtitle
        sub.font = Font(name="Aptos", size=11, italic=True, color=DARK)


def set_header(row):
    for cell in row:
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def format_range(ws, min_row, max_row, min_col, max_col, fill=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)


def widths(ws, mapping):
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


def freeze_and_filter(ws, row, end_col):
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(end_col)}{ws.max_row}"


def add_validation(ws, rng, values):
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def kW_for_hour(hour):
    if hour == 0:
        return 2.0
    if 1 <= hour <= 6:
        return 1.0
    if 7 <= hour <= 9:
        return 4.0
    if 10 <= hour <= 13:
        return 2.0
    if 14 <= hour <= 15:
        return 4.0
    if 16 <= hour <= 19:
        return 3.0
    return 2.0


def shift_for_hour(hour):
    if 7 <= hour <= 19:
        return "obsługa dzienna / ręczny nadzór"
    return "automatyczny zapis danych; bez obsługi nocnej"


def mode_for_hour(hour):
    if 8 <= hour <= 15:
        return "PV + BESS / falownik"
    if 16 <= hour <= 18:
        return "BESS + ewentualny agregat"
    if hour in (6, 7, 19):
        return "przejście / stabilizacja"
    return "BESS / tryb cichy"


def event_for_hour(hour):
    events = {
        7: "kontrola poranna, cold start, sprawdzenie logów nocnych",
        9: "pik 13 kW przez 3 min; zapis U/I/Hz/SOC przed-w trakcie-po",
        11: "test priorytetu PV/BESS; obserwacja ładowania",
        14: "test asymetrii faz L1/L2/L3 bez przekroczeń limitu",
        16: "symulacja niskiego SOC; start agregatu i ładowanie BESS",
        18: "E-STOP, restart kontrolowany, potwierdzenie braku samoczynnego startu",
    }
    return events.get(hour, "")


wb = Workbook()
ws = wb.active
ws.title = "00_Instrukcja"

style_title(ws, "A1", "Plan badań EcoGeneratora Power Protect 13 kW 3F", "Skoroszyt do dobowego testu pracy, FAT i rejestracji parametrów")
ws.merge_cells("A1:H1")
ws.merge_cells("A2:H2")
ws["A4"] = "Cel badania"
ws["B4"] = (
    "Sprawdzenie stabilności pracy hybrydowego generatora 13 kW pod profilem dobowym "
    "klienta oraz w kontrolowanych zdarzeniach: piki obciążenia, asymetria faz, niski SOC, "
    "start agregatu, ładowanie BESS, E-STOP, PV i praca bez ręcznej obsługi nocnej."
)
ws["A5"] = "Najważniejsza zasada"
ws["B5"] = (
    "Maksymalna moc badanego EcoGeneratora to 13 kW. Obciążnica może mieć większy zapas, "
    "ale nie jest to cel testu. Nocne godziny są tylko do automatycznego logowania; zdarzenia "
    "awaryjne i dolewanie paliwa planować w oknie dziennym."
)
ws["A7"] = "Kolejność pracy"
steps = [
    "Uzupełnij metryczkę i limity w arkuszu 01.",
    "Przejdź przez checklistę bezpieczeństwa w arkuszu 02.",
    "Zrealizuj profil dobowy z arkusza 03; kolumny Godzina doby i Obciążenie są bazą badania.",
    "Na bieżąco wpisuj pomiary w arkuszu 04, minimum co godzinę oraz przy zdarzeniach.",
    "Zapisuj paliwo, alarmy i temperatury szczególne w arkuszu 05.",
    "Wykonaj scenariusze FAT z arkuszy 06-09 tylko w godzinach obsługi dziennej.",
    "Podsumuj wynik w arkuszu 10."
]
for i, step in enumerate(steps, 8):
    ws.cell(i, 1).value = i - 7
    ws.cell(i, 2).value = step
format_range(ws, 4, 14, 1, 8)
ws["A17"] = "Źródła założeń"
sources = [
    "Badanie_generatora_13kW.docx: pomiary U, I, f, P, S, cosφ, temperatura, spadek napięcia, stopnie 0/25/50/75/100%.",
    "Scenariusze testów GG: tryby BESS/falownik/agregat/PV, asymetria, niski SOC, E-STOP, 24h profil pracy.",
    "Protokół FAT: struktura metryczki, checklist, rejestr próbek, alarmy, wnioski końcowe.",
    "Prospekt i prezentacje: EG 13 kW, PV ok. 2,46 kWp, BESS ok. 20 kWh, backup agregat 13 kW, monitoring 24/7, tryb cichy nocą."
]
for i, source in enumerate(sources, 18):
    ws.cell(i, 1).value = "•"
    ws.cell(i, 2).value = source
format_range(ws, 17, 21, 1, 8)
widths(ws, {"A": 18, "B": 95, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})
for row in range(4, 22):
    ws.row_dimensions[row].height = 34
if PHOTO.exists():
    img = Image(str(PHOTO))
    img.width = 360
    img.height = 260
    ws.add_image(img, "J4")

ws = wb.create_sheet("01_Metryczka_limity")
style_title(ws, "A1", "Metryczka urządzenia i limity badania")
ws.merge_cells("A1:H1")
data = [
    ("Nazwa urządzenia", "EcoGenerator Power Protect 13 kW 3F", "Numer seryjny", ""),
    ("Data badania", "", "Miejsce badania", ""),
    ("Osoba prowadząca", "", "Kwalifikacje / uprawnienia", ""),
    ("Wersja firmware falownika", "", "Wersja BMS / sterownika", ""),
    ("PV", "ok. 2,46 kWp / do uzupełnienia", "Magazyn energii", "ok. 20 kWh / do uzupełnienia"),
    ("Agregat backup", "13 kW / do uzupełnienia", "Paliwo", "benzyna / diesel / metanol / inne"),
    ("Obciążnica", "rezystancyjna, fazy osobno", "Limit testu", "13 kW EcoGeneratora"),
]
for r, row in enumerate(data, 3):
    ws.append(row)
set_header(ws[2])
ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "Pole", "Wartość", "Pole", "Wartość"
limits_start = 12
limits = [
    ("Moc znamionowa", 13, "kW", "Nie przekraczać bez zgody konstruktora"),
    ("25% mocy", "=ROUND($B$12*25%,2)", "kW", "ok. 3,25 kW"),
    ("50% mocy", "=ROUND($B$12*50%,2)", "kW", "ok. 6,50 kW"),
    ("75% mocy", "=ROUND($B$12*75%,2)", "kW", "ok. 9,75 kW"),
    ("100% mocy", "=ROUND($B$12,2)", "kW", "13 kW"),
    ("Moc na fazę przy 100%", "=ROUND($B$12/3,2)", "kW/faza", "obciążenie symetryczne"),
    ("Napięcie fazowe nominalne", 230, "V L-N", "kontrola każdej fazy"),
    ("Napięcie międzyfazowe nominalne", 400, "V L-L", "kontrola par faz"),
    ("Częstotliwość nominalna", 50, "Hz", "rejestrować min/max"),
    ("Prąd orientacyjny 3F cosφ=1", "=ROUND(13000/(SQRT(3)*400),1)", "A/faza", "ok. 18,8 A"),
    ("Prąd orientacyjny 3F cosφ=0,8", "=ROUND(13000/(SQRT(3)*400*0.8),1)", "A/faza", "ok. 23,5 A"),
]
ws.cell(limits_start - 1, 1).value = "Limity robocze i punkty odniesienia"
ws.cell(limits_start - 1, 1).font = Font(bold=True, color=WHITE)
ws.cell(limits_start - 1, 1).fill = PatternFill("solid", fgColor=BLUE)
ws.merge_cells(start_row=limits_start - 1, start_column=1, end_row=limits_start - 1, end_column=4)
ws.append(["Parametr", "Wartość", "Jednostka", "Uwagi"])
set_header(ws[limits_start])
for row in limits:
    ws.append(row)
format_range(ws, 2, ws.max_row, 1, 4)
widths(ws, {"A": 32, "B": 28, "C": 18, "D": 56})

ws = wb.create_sheet("02_Checklista")
style_title(ws, "A1", "Checklisty bezpieczeństwa i gotowości")
ws.merge_cells("A1:H1")
headers = ["Etap", "Nr", "Kontrola", "Wymaganie / sposób sprawdzenia", "OK", "NOK", "N/D", "Uwagi / dowód"]
ws.append(headers)
set_header(ws[2])
checks = [
    ("Przed startem", 1, "Dokumentacja", "DTR, schemat, BOM, nastawy falownika/BMS/agregatu dostępne na stanowisku"),
    ("Przed startem", 2, "Identyfikacja", "SN urządzenia, falownika, BESS, agregatu i firmware wpisane w metryczkę"),
    ("Przed startem", 3, "PE/uziemienie", "Ciągłość PE i ochrona przeciwporażeniowa sprawdzona wg procedury"),
    ("Przed startem", 4, "Okablowanie", "Przekroje, izolacja, wtyki, brak zwiniętych przewodów pod obciążeniem"),
    ("Przed startem", 5, "Obciążnica", "Wentylacja, wyrzut gorącego powietrza, stabilne ustawienie, brak uszkodzeń"),
    ("Przed startem", 6, "Agregat", "Paliwo, olej, brak wycieków, odprowadzenie spalin, dostępny E-STOP"),
    ("Przed startem", 7, "BESS/BMS", "SOC startowy, temperatury, komunikacja, brak alarmów aktywnych"),
    ("W trakcie", 8, "Rejestrator", "Analizator/logi zapisują U/I/P/Hz/SOC/PV/agregat/alarmy"),
    ("W trakcie", 9, "Termika", "Kontrola falownika, BESS, komory agregatu, spalin i przewodów"),
    ("W trakcie", 10, "Granice mocy", "Nie przekraczać 13 kW sumarycznie bez decyzji konstruktora"),
    ("Po teście", 11, "Wyłączenie", "Bezpieczne odłączenie obciążnicy, schłodzenie, zapis logów i zdjęć"),
]
for row in checks:
    ws.append(row)
format_range(ws, 2, ws.max_row, 1, 8)
add_validation(ws, "E3:G40", ["OK", "NOK", "N/D"])
widths(ws, {"A": 18, "B": 7, "C": 24, "D": 72, "E": 10, "F": 10, "G": 10, "H": 42})
freeze_and_filter(ws, 2, 8)

ws = wb.create_sheet("03_Plan_dobowy")
style_title(ws, "A1", "Dobowy profil obciążenia i zdarzeń")
ws.merge_cells("A1:L1")
headers = [
    "Godzina doby", "Przedział", "Obciążenie planowane [kW]", "Obciążenie [% 13 kW]",
    "Tryb oczekiwany", "Obsługa człowieka", "Zdarzenie testowe", "Pik [kW]",
    "Czas piku [min]", "PV oczek. [kW]", "Agregat oczek.", "Uwagi"
]
ws.append(headers)
set_header(ws[2])
for h in range(24):
    start = f"{h:02d}:00"
    end = f"{(h + 1) % 24:02d}:00"
    load = kW_for_hour(h)
    peak = 13 if h == 9 else ""
    peak_min = 3 if h == 9 else ""
    pv = 1.5 if 9 <= h <= 15 else 0
    agg = "ON testowo" if h == 16 else ("OFF / standby" if h not in (17, 18) else "wg SOC")
    ws.append([h, f"{start}-{end}", load, f"=C{ws.max_row+1}/13", mode_for_hour(h), shift_for_hour(h), event_for_hour(h), peak, peak_min, pv, agg, ""])
format_range(ws, 2, ws.max_row, 1, 12)
for cell in ws["D"][2:]:
    cell.number_format = "0%"
for row in range(3, ws.max_row + 1):
    if ws.cell(row, 6).value.startswith("automatyczny"):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=GREY)
    if ws.cell(row, 7).value:
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)
widths(ws, {"A": 12, "B": 15, "C": 20, "D": 18, "E": 26, "F": 34, "G": 58, "H": 12, "I": 14, "J": 14, "K": 16, "L": 32})
freeze_and_filter(ws, 2, 12)

chart = LineChart()
chart.title = "Planowane obciążenie dobowe"
chart.y_axis.title = "kW"
chart.x_axis.title = "Godzina"
data = Reference(ws, min_col=3, min_row=2, max_row=26)
cats = Reference(ws, min_col=1, min_row=3, max_row=26)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 20
ws.add_chart(chart, "N3")

ws = wb.create_sheet("04_Rejestr_pomiarow")
style_title(ws, "A1", "Rejestr pomiarów godzinowych i próbek dodatkowych")
ws.merge_cells("A1:AD1")
headers = [
    "Data", "Godzina doby", "Czas próbki", "Obciążenie plan [kW]", "Obciążenie rzeczyw. [kW]",
    "L1 kW", "L2 kW", "L3 kW", "P suma [kW]", "U L1-N [V]", "U L2-N [V]", "U L3-N [V]",
    "U L-L min [V]", "U L-L max [V]", "I L1 [A]", "I L2 [A]", "I L3 [A]", "Hz",
    "cosφ", "kVA", "THD U [%]", "SOC [%]", "PV [kW]", "Agregat", "Ładowanie BESS [kW]",
    "Paliwo dodane [l]", "Temp zewn. [°C]", "Temp komora agregatu [°C]", "Temp BESS [°C]",
    "Temp falownika [°C]", "Temp spalin [°C]", "Hałas [dB(A)]", "Alarmy / uwagi", "Wynik"
]
ws.append(headers)
set_header(ws[2])
for h in range(24):
    row = ws.max_row + 1
    ws.append(["", h, f"{h:02d}:00", f"='03_Plan_dobowy'!C{h+3}", "", "", "", "", f"=SUM(F{row}:H{row})", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
for _ in range(12):
    row = ws.max_row + 1
    ws.append(["", "", "", "", "", "", "", "", f"=SUM(F{row}:H{row})", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
format_range(ws, 2, ws.max_row, 1, 34)
add_validation(ws, "X3:X80", ["OFF", "ON", "START", "STOP", "AWARIA", "N/D"])
add_validation(ws, "AH3:AH80", ["OK", "NOK", "Powtórzyć", "N/D"])
for col in range(4, 32):
    for cell in ws.iter_cols(min_col=col, max_col=col, min_row=3, max_row=ws.max_row):
        for c in cell:
            c.number_format = "0.00"
ws.conditional_formatting.add("E3:E80", CellIsRule(operator="greaterThan", formula=["13"], fill=PatternFill("solid", fgColor=RED)))
ws.conditional_formatting.add("R3:R80", CellIsRule(operator="notBetween", formula=["49", "51"], fill=PatternFill("solid", fgColor=YELLOW)))
ws.conditional_formatting.add("J3:L80", CellIsRule(operator="notBetween", formula=["207", "253"], fill=PatternFill("solid", fgColor=YELLOW)))
widths(ws, {get_column_letter(i): 13 for i in range(1, 35)})
widths(ws, {"C": 12, "D": 18, "E": 18, "W": 12, "X": 14, "Y": 18, "Z": 16, "AA": 16, "AB": 22, "AC": 16, "AD": 18, "AE": 16, "AF": 14, "AG": 44, "AH": 14})
freeze_and_filter(ws, 2, 34)

ws = wb.create_sheet("05_Zdarzenia_paliwo")
style_title(ws, "A1", "Rejestr paliwa, alarmów, temperatur i interwencji")
ws.merge_cells("A1:L1")
headers = ["Data", "Godzina", "Typ wpisu", "Opis zdarzenia", "Obciążenie [kW]", "SOC [%]", "Paliwo przed [l]", "Paliwo dodane [l]", "Paliwo po [l]", "Temp / miejsce", "Osoba", "Dowód / plik / zdjęcie"]
ws.append(headers)
set_header(ws[2])
for _ in range(30):
    ws.append(["", "", "", "", "", "", "", "", "", "", "", ""])
format_range(ws, 2, ws.max_row, 1, 12)
add_validation(ws, "C3:C80", ["paliwo", "alarm", "temperatura", "E-STOP", "restart", "pomiar dodatkowy", "uwaga"])
widths(ws, {"A": 13, "B": 11, "C": 18, "D": 50, "E": 14, "F": 12, "G": 14, "H": 14, "I": 14, "J": 24, "K": 18, "L": 34})
freeze_and_filter(ws, 2, 12)

def add_scenario_sheet(name, title, rows):
    ws = wb.create_sheet(name)
    style_title(ws, "A1", title)
    ws.merge_cells("A1:K1")
    headers = ["Krok", "Godzina", "Czas trwania", "Tryb", "Obciążenie [kW]", "L1/L2/L3 [kW]", "Warunek początkowy", "Co wykonać", "Oczekiwana reakcja", "Wynik", "Uwagi"]
    ws.append(headers)
    set_header(ws[2])
    for row in rows:
        ws.append(row)
    format_range(ws, 2, ws.max_row, 1, 11)
    add_validation(ws, f"J3:J{ws.max_row}", ["OK", "NOK", "Powtórzyć", "N/D"])
    widths(ws, {"A": 8, "B": 12, "C": 14, "D": 24, "E": 16, "F": 18, "G": 28, "H": 54, "I": 52, "J": 14, "K": 34})
    freeze_and_filter(ws, 2, 11)


add_scenario_sheet("06_Doba_typowa", "Scenariusz A: profil dobowy klienta bez obsługi nocnej", [
    (1, "00:00-06:00", "6 h", "BESS / tryb cichy", "2→1", "sym.", "pełny zapis automatyczny", "Brak ręcznej obsługi; tylko logi i monitoring.", "Brak alarmów, stabilne U/f, brak pracy agregatu poza algorytmem.", "", ""),
    (2, "07:00", "30 min", "cold start / kontrola", 4, "sym.", "operator na miejscu", "Sprawdzić logi nocne, SOC, paliwo, temperatury, wentylację.", "Gotowość do testów dziennych.", "", ""),
    (3, "09:00", "3 min + 10 min obserwacji", "pik mocy", 13, "4,33/4,33/4,33", "stabilna praca 4 kW", "Wymusić pik 13 kW przez 3 minuty.", "Brak wyłączenia, akceptowalny spadek U/f, zapis reakcji.", "", ""),
    (4, "10:00-13:00", "4 h", "PV + BESS", 2, "sym.", "PV dostępne", "Obserwować priorytet PV i ładowanie BESS.", "Agregat nie startuje bez potrzeby, SOC zgodny z bilansem.", "", ""),
    (5, "16:00", "60 min", "niski SOC / agregat", 3, "sym.", "SOC zgodnie z procedurą", "Zasymulować próg startu agregatu lub wymusić tryb ładowania.", "Start agregatu, ładowanie BESS, odbiory mają priorytet.", "", ""),
    (6, "20:00-24:00", "4 h", "BESS / standby", 2, "sym.", "brak operatora", "Tylko automatyczne logowanie.", "Cicha praca, brak planowych interwencji nocnych.", "", ""),
])

add_scenario_sheet("07_Testy_FAT", "Scenariusz B: FAT obciążeniowy dzienny 25/50/75/100%", [
    (1, "08:00", "15 min", "bez obciążenia", 0, "0/0/0", "urządzenie gotowe", "Start falownika bez obciążenia, zapis U/f/temperatur.", "Stabilizacja bez alarmów.", "", ""),
    (2, "08:30", "30 min", "25%", 3.25, "1,08/1,08/1,08", "po stabilizacji", "Obciążenie symetryczne 25%.", "Stabilne napięcia i prądy.", "", ""),
    (3, "09:15", "30 min", "50%", 6.5, "2,17/2,17/2,17", "po kroku 2", "Obciążenie symetryczne 50%.", "Brak nadmiernego spadku U/f i wzrostu temperatur.", "", ""),
    (4, "10:00", "30 min", "75%", 9.75, "3,25/3,25/3,25", "po kroku 3", "Obciążenie symetryczne 75%.", "Reakcja falownika/BESS bez deratingu.", "", ""),
    (5, "10:45", "15 min", "100%", 13, "4,33/4,33/4,33", "zatwierdzone przez prowadzącego", "Obciążenie 100% bez przekraczania 13 kW.", "Osiąga moc znamionową bez alarmów krytycznych.", "", ""),
])

add_scenario_sheet("08_Asymetria_skoki", "Scenariusz C: asymetria faz i skoki obciążenia", [
    (1, "12:00", "15 min", "asymetria łagodna", 4, "2/1/1", "praca stabilna", "Ustawić nierówny rozkład faz.", "Falownik utrzymuje napięcia, brak alarmów przeciążenia fazy.", "", ""),
    (2, "12:30", "15 min", "asymetria budowa", 5.5, "4/1/0,5", "zgoda prowadzącego", "Profil kontenerów: L1 wysokie, L2 średnie, L3 niskie.", "System nie uszkadza prądnicy/agregatu, loguje asymetrię jeśli dotyczy.", "", ""),
    (3, "13:00", "3 min", "skok mocy", 13, "4,33/4,33/4,33", "bazowo 2-3 kW", "Nagły skok do 13 kW na 3 min.", "Czas powrotu U/f do stabilności zapisany.", "", ""),
    (4, "13:30", "15 min", "spadek obciążenia", 1, "0,33/0,33/0,33", "po skoku", "Szybki powrót do niskiego obciążenia.", "Brak oscylacji, brak błędów BMS/falownika.", "", ""),
])

add_scenario_sheet("09_Awaryjne", "Scenariusz D: zabezpieczenia i sytuacje awaryjne w oknie dziennym", [
    (1, "14:30", "30 min", "E-STOP", 2, "sym.", "operator przy urządzeniu", "Wcisnąć E-STOP zgodnie z procedurą.", "Bezpieczne odcięcie, brak samoczynnego restartu, alarm w logach.", "", ""),
    (2, "15:15", "30 min", "restart kontrolowany", 1, "sym.", "po E-STOP", "Wykonać restart według instrukcji.", "Powrót do pracy po potwierdzeniu przez operatora.", "", ""),
    (3, "16:00", "60 min", "symulacja końca paliwa", 3, "sym.", "tylko jeśli bezpieczne", "Zasymulować niski poziom paliwa czujnikiem/procedurą, bez ryzykownego opróżniania instalacji.", "System alarmuje i przechodzi w stan bezpieczny.", "", ""),
    (4, "17:15", "30 min", "utrata komunikacji", 2, "sym.", "procedura fault injection", "Zasymulować utratę BMS-falownik tylko zgodnie z procedurą konstruktora.", "Falownik/BMS przechodzi w stan bezpieczny, jasny alarm.", "", ""),
])

ws = wb.create_sheet("10_Podsumowanie")
style_title(ws, "A1", "Podsumowanie wyników i decyzja")
ws.merge_cells("A1:H1")
rows = [
    ("Obszar", "Ocena", "Najważniejsze obserwacje", "Działanie korygujące / decyzja"),
    ("Profil dobowy", "", "", ""),
    ("Stabilność napięcia i częstotliwości", "", "", ""),
    ("Praca z PV i magazynem energii", "", "", ""),
    ("Start i praca agregatu", "", "", ""),
    ("Piki 13 kW", "", "", ""),
    ("Asymetria faz", "", "", ""),
    ("E-STOP i restart", "", "", ""),
    ("Termika", "", "", ""),
    ("Paliwo i obsługa", "", "", ""),
    ("Alarmy / logi", "", "", ""),
]
for row in rows:
    ws.append(row)
set_header(ws[2])
format_range(ws, 2, ws.max_row, 1, 4)
add_validation(ws, "B3:B20", ["OK", "NOK", "Warunkowo", "Powtórzyć", "N/D"])
ws["A15"] = "Decyzja końcowa"
ws["B15"] = ""
ws["C15"] = "Podpis prowadzącego"
ws["D15"] = ""
ws["A17"] = "Uwagi końcowe"
ws.merge_cells("B17:D21")
format_range(ws, 15, 21, 1, 4)
widths(ws, {"A": 34, "B": 18, "C": 58, "D": 58})

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Aptos") if cell.font else Font(name="Aptos")
    for row in range(1, ws.max_row + 1):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 24

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(OUT)
